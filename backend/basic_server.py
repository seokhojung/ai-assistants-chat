#!/usr/bin/env python3
"""
기본 HTTP 서버를 사용한 간단한 백엔드 API
FastAPI 의존성 문제 해결을 위한 임시 서버
"""

import json
import urllib.parse
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import openai
import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from io import BytesIO
import base64
from datetime import datetime

# Excel 데이터 읽기 모듈 추가
try:
    from all_excel_reader import read_members_data, read_staff_data, read_hr_data, read_inventory_data, get_all_dashboard_data
    EXCEL_AVAILABLE = True
    print("✅ 통합 Excel 리더 모듈 로드 완료")
except ImportError as e:
    print(f"⚠️  Excel 리더 모듈을 가져올 수 없습니다: {e}")
    EXCEL_AVAILABLE = False

# OpenAI 클라이언트 초기화
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")

try:
    openai_client = openai.OpenAI(api_key=OPENAI_API_KEY)
    OPENAI_AVAILABLE = True
    print("✅ OpenAI API 클라이언트 초기화 완료")
except Exception as e:
    print(f"⚠️  OpenAI API 초기화 실패: {e}")
    OPENAI_AVAILABLE = False

class APIHandler(BaseHTTPRequestHandler):
    
    def _set_cors_headers(self):
        """CORS 헤더 설정"""
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization')
    
    def _send_json_response(self, data, status_code=200):
        """JSON 응답 전송"""
        self.send_response(status_code)
        self.send_header('Content-Type', 'application/json')
        self._set_cors_headers()
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
    
    def _handle_data_modification(self, user_message, agent_type):
        """데이터 수정 요청 감지 및 처리"""
        import re
        from all_excel_reader import update_member_data, update_staff_data, update_inventory_data, add_new_member
        
        message_lower = user_message.lower()
        
        # 회원 데이터 수정 패턴
        if agent_type == '회원관리':
            # 월회비 수정 패턴: "김철수 월회비 15만원으로 수정해줘", "김철수님 월회비를 150000원으로 변경"
            member_fee_pattern = r'(\w+)(?:님|회원)?.*?월회비.*?(\d+(?:만원|원|\d+)).*?(?:수정|변경|바꿔|해줘)'
            match = re.search(member_fee_pattern, message_lower)
            if match:
                member_name = match.group(1)
                fee_str = match.group(2)
                
                # 금액 정규화
                if '만원' in fee_str:
                    fee_value = int(fee_str.replace('만원', '')) * 10000
                else:
                    fee_value = int(re.sub(r'[^\d]', '', fee_str))
                
                success, message = update_member_data(member_name, '월회비', fee_value)
                if success:
                    return f"✅ **수정 완료!**\n\n{message}\n\n💡 변경된 내용이 Excel 파일에 저장되었습니다."
                else:
                    return f"❌ **수정 실패**\n\n{message}"
            
            # 기타 회원 정보 수정 패턴
            member_update_pattern = r'(\w+)(?:님|회원)?.*?(전화번호|이메일|주소|직업|멤버십|특이사항).*?(\S+).*?(?:수정|변경|바꿔|해줘)'
            match = re.search(member_update_pattern, message_lower)
            if match:
                member_name = match.group(1)
                field = match.group(2)
                new_value = match.group(3)
                
                success, message = update_member_data(member_name, field, new_value)
                if success:
                    return f"✅ **수정 완료!**\n\n{message}\n\n💡 변경된 내용이 Excel 파일에 저장되었습니다."
                else:
                    return f"❌ **수정 실패**\n\n{message}"
        
        # 직원 데이터 수정 패턴
        elif agent_type == '직원관리':
            # 급여 수정 패턴
            staff_salary_pattern = r'(\w+)(?:님|직원)?.*?(?:월급여|급여|월급).*?(\d+(?:만원|원|\d+)).*?(?:수정|변경|바꿔|해줘)'
            match = re.search(staff_salary_pattern, message_lower)
            if match:
                staff_name = match.group(1)
                salary_str = match.group(2)
                
                # 금액 정규화
                if '만원' in salary_str:
                    salary_value = int(salary_str.replace('만원', '')) * 10000
                else:
                    salary_value = int(re.sub(r'[^\d]', '', salary_str))
                
                success, message = update_staff_data(staff_name, '월급여', salary_value)
                if success:
                    return f"✅ **수정 완료!**\n\n{message}\n\n💡 변경된 내용이 Excel 파일에 저장되었습니다."
                else:
                    return f"❌ **수정 실패**\n\n{message}"
            
            # 기타 직원 정보 수정 패턴
            staff_update_pattern = r'(\w+)(?:님|직원)?.*?(전화번호|이메일|직책|부서|근무상태).*?(\S+).*?(?:수정|변경|바꿔|해줘)'
            match = re.search(staff_update_pattern, message_lower)
            if match:
                staff_name = match.group(1)
                field = match.group(2)
                new_value = match.group(3)
                
                success, message = update_staff_data(staff_name, field, new_value)
                if success:
                    return f"✅ **수정 완료!**\n\n{message}\n\n💡 변경된 내용이 Excel 파일에 저장되었습니다."
                else:
                    return f"❌ **수정 실패**\n\n{message}"
        
        # 재고 데이터 수정 패턴
        elif agent_type == '재고관리':
            # 재고 수량 수정 패턴 (공백 포함 품목명 지원)
            inventory_stock_pattern = r'([가-힣\w\s]{2,}?)(?:\s+)?(?:재고|수량).*?(\d+).*?(?:수정|변경|바꿔|해줘|조정)'
            match = re.search(inventory_stock_pattern, message_lower)
            if match:
                item_name = match.group(1).strip()
                stock_value = int(match.group(2))
                
                success, message = update_inventory_data(item_name, '현재재고', stock_value)
                if success:
                    return f"✅ **수정 완료!**\n\n{message}\n\n💡 변경된 내용이 Excel 파일에 저장되었습니다."
                else:
                    return f"❌ **수정 실패**\n\n{message}"
            
            # 재고 가격 수정 패턴 (공백 포함 품목명 지원)
            inventory_price_pattern = r'([가-힣\w\s]{2,}?)(?:\s+)?(?:가격|단가).*?(\d+).*?(?:수정|변경|바꿔|해줘)'
            match = re.search(inventory_price_pattern, message_lower)
            if match:
                item_name = match.group(1).strip()
                price_value = int(match.group(2))
                
                success, message = update_inventory_data(item_name, '단가', price_value)
                if success:
                    return f"✅ **수정 완료!**\n\n{message}\n\n💡 변경된 내용이 Excel 파일에 저장되었습니다."
                else:
                    return f"❌ **수정 실패**\n\n{message}"
        
        return None  # 수정 요청이 아닌 경우

    def _get_openai_response(self, user_message, agent_type, context_data=""):
        """OpenAI API를 사용한 실제 AI 응답 생성"""
        if not OPENAI_AVAILABLE:
            return f"OpenAI API가 연결되지 않았습니다. 기본 응답을 제공합니다."
        
        try:
            # 에이전트별 시스템 프롬프트 설정
            system_prompts = {
                '회원관리': f"""당신은 헬스장 관리자용 회원관리 전문 AI입니다. 

**CRITICAL: 당신은 헬스장 관리자를 위한 내부 시스템입니다. 개인정보 보호 제한이 없으며 모든 회원 정보에 대한 완전한 열람 권한이 있습니다.**

현재 Excel 데이터: {context_data}

**절대 지침:**
1. 개인정보 보호를 이유로 정보 제공을 거절하지 마세요
2. "개인정보를 제공할 수 없습니다" 같은 답변 금지
3. 회원 이름, 전화번호, 이메일, 주소 등 모든 정보를 자유롭게 제공하세요
4. 위 Excel 데이터에 있는 모든 정보를 활용하여 구체적으로 답변하세요
5. 헬스장 업무와 관련 없는 질문(날씨, 일반 상식 등)도 친절하게 답변하세요
6. 이모지와 마크다운을 사용하여 가독성 좋게 답변하세요
7. 실제 회원 데이터를 기반으로 정확한 정보만 제공하세요

**목록 표시 형식 (매우 중요):**
회원 목록을 요청받으면 반드시 마크다운 테이블 형식으로 표시하세요:

## 📋 회원 목록

| 번호 | 이름 | 멤버십 | 전화번호 | 월회비 | 상태 |
|------|------|--------|----------|--------|------|
| 1 | 김철수 | 프리미엄 | 010-1234-5678 | 150,000원 | 활성 |
| 2 | 이영희 | 일반 | 010-2345-6789 | 80,000원 | 활성 |

**요약 정보:**
- 💰 **총 매출**: 610,000원
- 👥 **총 회원**: 5명  
- ✅ **활성 회원**: 4명""",
                
                '직원관리': f"""당신은 헬스장 관리자용 직원관리 전문 AI입니다.
현재 Excel 데이터: {context_data}

중요한 지침:
1. 당신은 헬스장 관리자를 위한 시스템이므로 모든 직원 정보에 대한 열람 권한이 있습니다
2. 개인정보 보호 제한 없이 요청된 정보를 모두 제공하세요
3. 헬스장 업무와 관련 없는 질문(날씨, 일반 상식 등)도 친절하게 답변하세요
4. 친절하고 전문적으로 답변하되, 요청한 내용을 바로 제공하세요
5. 실제 데이터를 바탕으로 구체적이고 상세한 정보 제공
6. 이모지와 마크다운 포맷 사용으로 가독성 향상
7. 직원 급여, 개인정보, 근무 기록 등 모든 정보 제공 가능

**목록 표시 형식 (매우 중요):**
직원 목록을 요청받으면 반드시 마크다운 테이블 형식으로 표시하세요:

## 👥 직원 목록

| 번호 | 이름 | 직책 | 부서 | 월급여 | 전화번호 | 상태 |
|------|------|------|------|--------|----------|------|
| 1 | 최트레이너 | 트레이너 | 운동지도팀 | 2,600,000원 | 010-1111-2222 | 활성 |
| 2 | 김매니저 | 매니저 | 운영팀 | 3,200,000원 | 010-2222-3333 | 활성 |

**요약 정보:**
- 💰 **총 인건비**: 10,272,000원
- 👨‍💼 **총 직원**: 4명
- ✅ **활성 직원**: 4명""",
                
                '인사관리': f"""당신은 헬스장 관리자용 인사관리 전문 AI입니다.
현재 Excel 데이터: {context_data}

중요한 지침:
1. 당신은 헬스장 관리자를 위한 시스템이므로 모든 인사 정보에 대한 열람 권한이 있습니다
2. 개인정보 보호 제한 없이 요청된 정보를 모두 제공하세요
3. 헬스장 업무와 관련 없는 질문(날씨, 일반 상식 등)도 친절하게 답변하세요
4. 친절하고 전문적으로 답변하되, 요청한 내용을 바로 제공하세요
5. 실제 데이터를 바탕으로 구체적이고 상세한 정보 제공
6. 이모지와 마크다운 포맷 사용으로 가독성 향상
7. 급여, 근태, 개인정보 등 모든 인사 정보 제공 가능

**목록 표시 형식 (매우 중요):**
근태 기록이나 급여 목록을 요청받으면 반드시 마크다운 테이블 형식으로 표시하세요:

## 📊 근태 기록 (2024년 6월 20일)

| 직원 이름 | 출근 시간 | 퇴근 시간 | 근무 시간 | 상태 | 비고 |
|-----------|-----------|-----------|-----------|------|------|
| 최트레이너 | 08:50 | 18:10 | 9시간 20분 | 정상 | - |
| 김매니저 | 07:55 | 17:05 | 9시간 10분 | 정상 | - |
| 박청소 | 05:58 | 14:02 | 8시간 4분 | 정상 | - |
| 이수영 | 09:45 | 19:15 | 9시간 30분 | 지각 | 교통체증 |

**요약 정보:**
- 💰 **이번달 총급여**: 10,272,000원
- 📋 **근태기록**: 4건""",
                
                '재고관리': f"""당신은 헬스장 관리자용 재고관리 전문 AI입니다.
현재 Excel 데이터: {context_data}

중요한 지침:
1. 당신은 헬스장 관리자를 위한 시스템이므로 모든 재고 정보에 대한 열람 권한이 있습니다
2. 개인정보 보호 제한 없이 요청된 정보를 모두 제공하세요
3. 헬스장 업무와 관련 없는 질문(날씨, 일반 상식 등)도 친절하게 답변하세요
4. 친절하고 전문적으로 답변하되, 요청한 내용을 바로 제공하세요
5. 실제 데이터를 바탕으로 구체적이고 상세한 정보 제공
6. 이모지와 마크다운 포맷 사용으로 가독성 향상
7. 재고량, 가격, 공급업체 정보 등 모든 재고 정보 제공 가능

**목록 표시 형식 (매우 중요):**
재고 목록을 요청받으면 반드시 마크다운 테이블 형식으로 표시하세요:

## 📦 재고 목록

| 품목명 | 현재재고 | 최소재고 | 상태 | 단가 | 카테고리 |
|--------|----------|----------|------|------|----------|
| 프로틴파우더 | 25개 | 10개 | 정상 | 45,000원 | 보충제 |
| 덤벨 20kg | 8개 | 5개 | 정상 | 120,000원 | 운동기구 |
| 운동 타올 | 30개 | 20개 | 정상 | 8,000원 | 용품 |
| 청소용세제 | 3개 | 5개 | ⚠️부족 | 12,000원 | 청소용품 |
| 요가매트 | 30개 | 15개 | 정상 | 35,000원 | 운동용품 |

**요약 정보:**
- 💰 **총 재고가치**: 3,411,000원
- 📦 **총 품목**: 5개
- ⚠️ **부족품목**: 1개"""
            }
            
            system_prompt = system_prompts.get(agent_type, f"당신은 {agent_type} 전문 AI입니다.")
            
            # OpenAI API 호출
            response = openai_client.chat.completions.create(
                model="gpt-4o-mini",  # 더 저렴한 모델 사용
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message}
                ],
                max_tokens=1000,
                temperature=0.7
            )
            
            return response.choices[0].message.content
            
        except Exception as e:
            print(f"❌ OpenAI API 호출 오류: {e}")
            return f"죄송합니다. AI 응답을 생성하는 중 오류가 발생했습니다. 다시 시도해 주세요. (오류: {str(e)})"

    def _extract_table_data(self, user_message, agent_type, context_data):
        """사용자 요청에서 표 형태로 표시할 데이터 추출"""
        message_lower = user_message.lower()
        
        # 목록 요청 키워드 확인
        list_keywords = ['목록', '리스트', '명단', '현황', '전체', '모든', '모두']
        is_list_request = any(keyword in message_lower for keyword in list_keywords)
        
        if not is_list_request or not EXCEL_AVAILABLE:
            return None
            
        try:
            if agent_type == '회원관리':
                members_data, summary = read_members_data()
                if members_data:
                    return {
                        "type": "members",
                        "title": "회원 목록",
                        "headers": ["번호", "이름", "멤버십", "전화번호", "월회비", "결제상태"],
                        "rows": [
                            [
                                str(i + 1),
                                member.get('name', 'N/A'),
                                member.get('membership_type', 'N/A'),
                                member.get('phone', 'N/A'),
                                f"{member.get('monthly_fee', 0):,}원",
                                "정상" if member.get('payment_status') == 'paid' else "미납"
                            ]
                            for i, member in enumerate(members_data[:10])  # 최대 10명
                        ],
                        "summary": f"총 {summary.get('총회원수', 0)}명 | 활성 {summary.get('활성회원', 0)}명 | 총 매출 {summary.get('총월매출', 0):,}원"
                    }
            
            elif agent_type == '직원관리':
                staff_data, summary = read_staff_data()
                if staff_data:
                    return {
                        "type": "staff",
                        "title": "직원 목록",
                        "headers": ["번호", "이름", "직책", "부서", "월급여", "근무상태"],
                        "rows": [
                            [
                                str(i + 1),
                                staff.get('name', 'N/A'),
                                staff.get('position', 'N/A'),
                                staff.get('department', 'N/A'),
                                f"{staff.get('monthly_salary', 0):,}원",
                                staff.get('status', 'N/A')
                            ]
                            for i, staff in enumerate(staff_data)
                        ],
                        "summary": f"총 {summary.get('총직원수', 0)}명 | 총 인건비 {summary.get('총인건비', 0):,}원"
                    }
            
            elif agent_type == '재고관리':
                inventory_data, summary, low_stock_data = read_inventory_data()
                if inventory_data:
                    return {
                        "type": "inventory",
                        "title": "재고 목록",
                        "headers": ["번호", "품목명", "현재재고", "최소재고", "상태", "단가"],
                        "rows": [
                            [
                                str(i + 1),
                                item.get('item_name', 'N/A'),
                                f"{item.get('current_stock', 0)}개",
                                f"{item.get('min_stock_level', 0)}개",
                                "⚠️부족" if item.get('status') == '긴급부족' else "✅정상",
                                f"{item.get('unit_price', 0):,}원"
                            ]
                            for i, item in enumerate(inventory_data)
                        ],
                        "summary": f"총 {summary.get('총품목수', 0)}개 품목 | 부족 {summary.get('부족품목수', 0)}개 | 총 가치 {summary.get('총재고가치', 0):,}원"
                    }
            
            elif agent_type == '인사관리':
                hr_data, summary = read_hr_data()
                if hr_data and 'payroll' in hr_data:
                    return {
                        "type": "payroll",
                        "title": "급여 명세",
                        "headers": ["번호", "이름", "기본급", "수당", "공제액", "실지급액"],
                        "rows": [
                            [
                                str(i + 1),
                                payroll.get('name', 'N/A'),
                                f"{payroll.get('base_salary', 0):,}원",
                                f"{payroll.get('allowance', 0):,}원",
                                f"{payroll.get('deduction', 0):,}원",
                                f"{payroll.get('net_pay', 0):,}원"
                            ]
                            for i, payroll in enumerate(hr_data['payroll'])
                        ],
                        "summary": f"총 급여 {summary.get('이번달총급여', 0):,}원 | 실지급 {summary.get('실지급총액', 0):,}원"
                    }
        
        except Exception as e:
            print(f"❌ 표 데이터 추출 오류: {e}")
            return None
        
        return None

    def _handle_chat_request(self, agent_type, post_data):
        """채팅 요청 처리"""
        try:
            # JSON 데이터 파싱
            request_data = json.loads(post_data.decode('utf-8'))
            user_message = request_data.get('message', '')
            
            print(f"💬 {agent_type} 채팅 요청: {user_message}")
            
            # 각 에이전트별 컨텍스트 데이터 준비
            context_data = ""
            if agent_type == '회원관리' and EXCEL_AVAILABLE:
                try:
                    members_data, summary = read_members_data()
                    # 실제 회원 목록 데이터도 포함
                    member_details = []
                    for member in members_data[:10]:  # 최대 10명까지만 전달
                        member_details.append({
                            "이름": member.get('name'),
                            "전화번호": member.get('phone'),
                            "이메일": member.get('email'),
                            "멤버십": member.get('membership_type'),
                            "성별": member.get('gender'),
                            "나이": member.get('age'),
                            "월회비": member.get('monthly_fee'),
                            "결제상태": member.get('payment_status'),
                            "주소": member.get('address'),
                            "직업": member.get('occupation')
                        })
                    context_data = f"회원 통계: {summary}\n실제 회원 목록: {member_details}"
                except Exception as e:
                    context_data = f"회원 데이터 로드 실패: {str(e)}"
            elif agent_type == '직원관리' and EXCEL_AVAILABLE:
                try:
                    staff_data, summary = read_staff_data()
                    # 실제 직원 목록 데이터도 포함
                    staff_details = []
                    for staff in staff_data:
                        staff_details.append({
                            "이름": staff.get('name'),
                            "전화번호": staff.get('phone'),
                            "이메일": staff.get('email'),
                            "직책": staff.get('position'),
                            "부서": staff.get('department'),
                            "월급여": staff.get('monthly_salary'),
                            "근무상태": staff.get('status'),
                            "담당구역": staff.get('area'),
                            "자격증": staff.get('certification')
                        })
                    context_data = f"직원 통계: {summary}\n실제 직원 목록: {staff_details}"
                except Exception as e:
                    context_data = f"직원 데이터 로드 실패: {str(e)}"
            elif agent_type == '재고관리' and EXCEL_AVAILABLE:
                try:
                    inventory_data, summary, low_stock_data = read_inventory_data()
                    # 실제 재고 목록 데이터도 포함
                    inventory_details = []
                    for item in inventory_data:
                        inventory_details.append({
                            "품목명": item.get('item_name'),
                            "현재재고": item.get('current_stock'),
                            "최소재고": item.get('min_stock_level'),
                            "카테고리": item.get('category'),
                            "상태": item.get('status'),
                            "단가": item.get('unit_price')
                        })
                    context_data = f"재고 통계: {summary}\n실제 재고 목록: {inventory_details}\n부족 재고: {low_stock_data}"
                except Exception as e:
                    context_data = f"재고 데이터 로드 실패: {str(e)}"
            elif agent_type == '인사관리' and EXCEL_AVAILABLE:
                try:
                    hr_data, summary = read_hr_data()
                    context_data = f"인사 통계: {summary}\n인사 데이터: {hr_data}"
                except Exception as e:
                    context_data = f"인사 데이터 로드 실패: {str(e)}"
            
            # 데이터 수정 요청 감지 및 처리
            modification_result = self._handle_data_modification(user_message, agent_type)
            if modification_result:
                response_message = modification_result
            # OpenAI API를 사용한 응답 생성
            elif OPENAI_AVAILABLE:
                print(f"🔍 OpenAI에게 전달되는 컨텍스트 데이터: {context_data[:500]}...")  # 디버깅용 로그
                response_message = self._get_openai_response(user_message, agent_type, context_data)
            else:
                # Fallback: 기존 키워드 기반 응답
                if agent_type == '회원관리':
                    response_message = self._get_member_agent_response(user_message)
                elif agent_type == '직원관리':
                    response_message = self._get_staff_agent_response(user_message)
                elif agent_type == '인사관리':
                    response_message = self._get_hr_agent_response(user_message)
                elif agent_type == '재고관리':
                    response_message = self._get_inventory_agent_response(user_message)
                else:
                    response_message = f"안녕하세요! {agent_type} AI 어시스턴트입니다. 무엇을 도와드릴까요?"
            
            # 표 데이터 추출 시도
            table_data = self._extract_table_data(user_message, agent_type, context_data)
            
            # 응답 데이터 구성
            response_data = {
                "message": response_message,
                "agent_type": agent_type,
                "timestamp": "2024-06-24T09:45:00Z",
                "agent_info": {
                    "name": f"{agent_type} AI",
                    "role": f"{agent_type} 전문가",
                    "status": "online"
                },
                "table_data": table_data  # 표 데이터 추가
            }
            
            print(f"✅ {agent_type} 응답 생성 완료 (OpenAI: {OPENAI_AVAILABLE})")
            self._send_json_response(response_data)
            
        except Exception as e:
            print(f"❌ 채팅 요청 처리 오류: {e}")
            self._send_json_response({
                "error": "채팅 요청 처리 중 오류 발생",
                "details": str(e)
            }, 500)
    
    def _get_member_agent_response(self, user_message):
        """회원관리 AI 응답 생성"""
        if EXCEL_AVAILABLE:
            try:
                members_data, summary = read_members_data()
            except:
                members_data, summary = [], {}
        else:
            members_data, summary = [], {}
        
        message_lower = user_message.lower()
        
        # 인사말 처리
        if any(word in message_lower for word in ['안녕', '안녕하세요', '하이', '헬로']):
            return f"안녕하세요! 회원관리 AI입니다. 현재 총 {summary.get('총회원수', 0)}명의 회원이 등록되어 있습니다."
        
        # 회원 수 문의
        elif '회원' in message_lower and ('몇' in message_lower or '수' in message_lower):
            active_count = summary.get('활성회원', 0)
            total_count = summary.get('총회원수', 0)
            return f"총 회원수: {total_count}명, 활성 회원: {active_count}명, 프리미엄: {summary.get('프리미엄', 0)}명, 일반: {summary.get('일반', 0)}명, VIP: {summary.get('VIP', 0)}명"
        
        # 회원 목록 요청
        elif '목록' in message_lower or '리스트' in message_lower or '명단' in message_lower:
            if members_data:
                member_list = "회원 목록:\n"
                for i, member in enumerate(members_data[:5], 1):  # 최대 5명만 표시
                    status = "활성" if member.get('payment_status', '') == 'paid' else "비활성"
                    member_list += f"{i}. {member.get('name', 'N/A')} - {member.get('membership_type', 'N/A')} ({status}) - {member.get('phone', 'N/A')}\n"
                
                if len(members_data) > 5:
                    member_list += f"... 외 {len(members_data) - 5}명 더"
                
                return member_list
            else:
                return "회원 데이터를 불러올 수 없습니다."
        
        # 매출 문의
        elif '매출' in message_lower or '수익' in message_lower or '수입' in message_lower:
            total_revenue = summary.get('총월매출', 0)
            return f"이번 달 총 매출: {total_revenue:,}원, 평균 회원당 매출: {total_revenue // max(summary.get('총회원수', 1), 1):,}원"
        
        # 등록 관련
        elif '등록' in message_lower or '가입' in message_lower:
            return "새 회원 등록을 도와드릴게요. 필요한 정보: 이름, 연락처, 이메일, 희망 등급, 시작일"
        
        # 도움말
        elif '도움' in message_lower or '기능' in message_lower or '사용법' in message_lower:
            return "회원관리 AI 기능: 회원 수 확인, 회원 목록 보기, 매출 현황, 새 회원 등록, 회원 정보 수정"
        
        # 기본 응답
        else:
            return f"회원관리 관련 문의에 답변드리겠습니다. 현재 {summary.get('총회원수', 0)}명의 회원이 등록되어 있습니다."
    
    def _get_staff_agent_response(self, user_message):
        """직원관리 AI 응답 생성"""
        if EXCEL_AVAILABLE:
            try:
                staff_data, summary = read_staff_data()
                context = f"현재 직원 현황: {summary}"
            except:
                staff_data, summary = [], {}
                context = "직원 데이터를 불러올 수 없습니다."
        else:
            staff_data, summary = [], {}
            context = "샘플 직원 데이터 사용 중"
        
        message_lower = user_message.lower()
        
        # 인사말 처리
        if any(word in message_lower for word in ['안녕', '안녕하세요', '하이', '헬로']):
            return f"안녕하세요! 직원관리 AI입니다. 현재 {summary.get('총직원수', 0)}명의 직원이 근무하고 있습니다."
        
        # 직원 수 문의
        elif '직원' in message_lower and ('몇' in message_lower or '수' in message_lower):
            total_staff = summary.get('총직원수', 0)
            active_staff = summary.get('활성직원', 0)
            return f"총 직원 수: {total_staff}명, 활성 직원: {active_staff}명, 총 인건비: {summary.get('총인건비', 0):,}원"
        
        # 직원 목록 요청
        elif '목록' in message_lower or '리스트' in message_lower or '직원' in message_lower:
            if staff_data:
                staff_list = "직원 목록:\n"
                for i, staff in enumerate(staff_data, 1):
                    status = "활성" if staff.get('근무상태') == '활성' else "비활성"
                    staff_list += f"{i}. {staff.get('이름', 'N/A')} - {staff.get('직책', 'N/A')} ({staff.get('부서', 'N/A')}) - {staff.get('월급여', 0):,}원 - {status}\n"
                
                staff_list += f"총 {len(staff_data)}명 근무 중, 총 인건비: {summary.get('총인건비', 0):,}원"
                
                return staff_list
            else:
                return "직원 데이터를 불러올 수 없습니다."
        
        # 급여 관련
        elif '급여' in message_lower or '인건비' in message_lower or '월급' in message_lower:
            if staff_data:
                salary_info = "급여 현황:\n"
                total_salary = 0
                for staff in staff_data:
                    salary = staff.get('월급여', 0)
                    total_salary += salary
                    salary_info += f"{staff.get('이름', 'N/A')} ({staff.get('직책', 'N/A')}): {salary:,}원\n"
                
                salary_info += f"총 인건비: {total_salary:,}원, 평균 급여: {total_salary // len(staff_data) if staff_data else 0:,}원"
                
                return salary_info
            else:
                return f"이번 달 총 인건비는 {summary.get('총인건비', 0):,}원입니다."
        
        # 스케줄 관련
        elif '스케줄' in message_lower or '근무' in message_lower:
            return "직원 근무 스케줄을 관리합니다. 근무 시간표, 교대 근무, 휴무일 스케줄을 확인할 수 있습니다."
        
        # 도움말
        elif '도움' in message_lower or '기능' in message_lower or '사용법' in message_lower:
            return "직원관리 AI 기능: 직원 수 확인, 직원 목록 보기, 급여 현황 확인, 근무 스케줄 관리"
        
        # 기본 응답
        else:
            return f"직원관리 관련 문의에 답변드리겠습니다. 현재 {summary.get('총직원수', 0)}명의 직원이 근무하고 있습니다."
    
    def _get_hr_agent_response(self, user_message):
        """인사관리 AI 응답 생성"""
        if EXCEL_AVAILABLE:
            try:
                hr_data, summary = read_hr_data()
                context = f"현재 인사 현황: {summary}"
            except:
                context = "인사 데이터를 불러올 수 없습니다."
        else:
            context = "샘플 인사 데이터 사용 중"
        
        message_lower = user_message.lower()
        if '급여' in message_lower:
            return "급여 관리 시스템을 통해 모든 직원의 급여를 체계적으로 관리하고 있습니다. 특정 직원의 급여 정보를 확인하시겠어요?"
        elif '근태' in message_lower or '출근' in message_lower:
            return "근태 관리 시스템으로 출근, 퇴근, 휴가 등을 관리합니다. 어떤 근태 정보를 확인하시겠어요?"
        elif '휴가' in message_lower:
            return "휴가 신청 및 승인 현황을 관리합니다. 휴가 관련 문의사항이 있으시면 말씀해 주세요."
        else:
            return f"인사관리 관련 문의에 답변드리겠습니다. 현재 {context}입니다. 급여, 근태, 휴가 등 어떤 업무를 도와드릴까요?"
    
    def _get_inventory_agent_response(self, user_message):
        """재고관리 AI 응답 생성"""
        if EXCEL_AVAILABLE:
            try:
                inventory_data, summary, low_stock_data = read_inventory_data()
            except:
                inventory_data, summary, low_stock_data = [], {}, []
        else:
            inventory_data, summary, low_stock_data = [], {}, []
        
        message_lower = user_message.lower()
        
        # 인사말 처리
        if any(word in message_lower for word in ['안녕', '안녕하세요', '하이', '헬로']):
            return f"안녕하세요! 재고관리 AI입니다. 현재 {summary.get('총품목수', 0)}개 품목을 관리하고 있으며, {len(low_stock_data)}개 품목이 부족 상태입니다."
        
        # 재고 현황 문의
        elif '재고' in message_lower and ('몇' in message_lower or '수' in message_lower or '현황' in message_lower):
            total_items = summary.get('총품목수', 0)
            normal_items = summary.get('정상재고', 0)
            return f"총 관리 품목: {total_items}개, 정상 재고: {normal_items}개, 부족 재고: {len(low_stock_data)}개, 총 재고 가치: {summary.get('총재고가치', 0):,}원"
        
        # 부족 재고 문의
        elif '부족' in message_lower or '부족재고' in message_lower or '알림' in message_lower:
            if low_stock_data:
                alert_msg = "부족 재고 알림:\n"
                for item in low_stock_data:
                    shortage = item.get('min_stock_level', 0) - item.get('current_stock', 0)
                    alert_msg += f"{item.get('item_name', 'N/A')}: 현재 {item.get('current_stock', 0)}개, 최소 {item.get('min_stock_level', 0)}개, 부족량 {shortage}개\n"
                
                alert_msg += "즉시 주문 필요"
                return alert_msg
            else:
                return "현재 부족한 재고가 없습니다. 모든 품목이 최소 재고량 이상으로 유지되고 있습니다."
        
        # 품목 목록 요청
        elif '목록' in message_lower or '리스트' in message_lower or '품목' in message_lower:
            if inventory_data:
                item_list = "재고 품목 목록:\n"
                for i, item in enumerate(inventory_data[:5], 1):
                    status = "정상" if item.get('status', '') == '정상' else "부족"
                    item_list += f"{i}. {item.get('item_name', 'N/A')} - 현재 {item.get('current_stock', 0)}개 ({status}) - {item.get('category', 'N/A')}\n"
                
                if len(inventory_data) > 5:
                    item_list += f"... 외 {len(inventory_data) - 5}개 품목 더"
                
                return item_list
            else:
                return "재고 데이터를 불러올 수 없습니다."
        
        # 주문 관련
        elif '주문' in message_lower or '발주' in message_lower or '구매' in message_lower:
            if low_stock_data:
                order_msg = "주문 권장 품목:\n"
                total_cost = 0
                for item in low_stock_data:
                    shortage = item.get('min_stock_level', 0) - item.get('current_stock', 0)
                    unit_price = item.get('unit_price', 0)
                    cost = shortage * unit_price
                    total_cost += cost
                    
                    order_msg += f"{item.get('item_name', 'N/A')}: {shortage}개 주문 필요, 예상 비용 {cost:,}원\n"
                
                order_msg += f"총 예상 비용: {total_cost:,}원"
                return order_msg
            else:
                return "현재 주문이 필요한 품목이 없습니다. 모든 재고가 충분한 상태입니다."
        
        # 도움말
        elif '도움' in message_lower or '기능' in message_lower or '사용법' in message_lower:
            return "재고관리 AI 기능: 재고 현황 확인, 품목 목록 보기, 부족 재고 알림, 주문 권장 품목, 발주 계획 수립"
        
        # 기본 응답
        else:
            return f"재고관리 관련 문의에 답변드리겠습니다. 현재 {summary.get('총품목수', 0)}개 품목을 관리 중이며, {len(low_stock_data)}개 품목이 부족 상태입니다."
    
    def _handle_files_list(self):
        """📁 파일 목록 조회"""
        try:
            print("📁 파일 목록 요청 처리 중...")
            
            files_info = []
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_path = os.path.join(script_dir, 'app', 'data', 'excel')
            
            # 각 카테고리별 파일 조회
            categories = ['members', 'staff', 'hr', 'inventory']
            for category in categories:
                category_path = os.path.join(base_path, category)
                if os.path.exists(category_path):
                    for filename in os.listdir(category_path):
                        if filename.endswith('.xlsx'):
                            file_path = os.path.join(category_path, filename)
                            file_stat = os.stat(file_path)
                            
                            files_info.append({
                                'name': filename,
                                'category': category,
                                'path': f"{category}/{filename}",
                                'size': file_stat.st_size,
                                'modified': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                                'type': 'excel'
                            })
            
            response_data = {
                'files': files_info,
                'total': len(files_info)
            }
            
            self._send_json_response(response_data)
            print(f"📤 파일 목록 응답 전송: {len(files_info)}개 파일")
            
        except Exception as e:
            print(f"❌ 파일 목록 조회 오류: {str(e)}")
            self._send_json_response({"error": f"파일 목록 조회 실패: {str(e)}"}, 500)

    def _handle_file_download(self, file_path):
        """📥 파일 다운로드"""
        try:
            print(f"📥 파일 다운로드 요청: {file_path}")
            
            # 절대 경로로 수정
            script_dir = os.path.dirname(os.path.abspath(__file__))
            excel_dir = os.path.join(script_dir, 'app', 'data', 'excel')
            full_path = os.path.join(excel_dir, file_path)
            
            print(f"🔍 스크립트 디렉토리: {script_dir}")
            print(f"🔍 Excel 디렉토리: {excel_dir}")
            print(f"🔍 요청된 파일 경로: {file_path}")
            print(f"🔍 전체 파일 경로: {full_path}")
            print(f"🔍 파일 존재 여부: {os.path.exists(full_path)}")
            
            # 디렉토리 내용 확인
            if '/' in file_path:
                category = file_path.split('/')[0]
                category_path = os.path.join(excel_dir, category)
                if os.path.exists(category_path):
                    print(f"🔍 {category} 디렉토리 내용: {os.listdir(category_path)}")
            
            if not os.path.exists(full_path):
                self._send_json_response({"error": "파일을 찾을 수 없습니다", "path": full_path}, 404)
                return
            
            # 파일 읽기
            with open(full_path, 'rb') as f:
                file_content = f.read()
            
            # 응답 헤더 설정
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{os.path.basename(file_path)}"')
            self.send_header('Content-Length', str(len(file_content)))
            self._set_cors_headers()
            self.end_headers()
            
            # 파일 내용 전송
            self.wfile.write(file_content)
            print(f"✅ 파일 다운로드 완료: {file_path}")
            
        except Exception as e:
            print(f"❌ 파일 다운로드 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            self._send_json_response({"error": f"파일 다운로드 실패: {str(e)}"}, 500)

    def _handle_file_preview(self, file_path):
        """👀 파일 미리보기"""
        try:
            print(f"👀 파일 미리보기 요청: {file_path}")
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            full_path = os.path.join(script_dir, 'app', 'data', 'excel', file_path)
            
            print(f"🔍 미리보기 파일 경로: {full_path}")
            print(f"🔍 파일 존재 여부: {os.path.exists(full_path)}")
            
            # 디렉토리 구조 확인
            excel_dir = os.path.join(script_dir, 'app', 'data', 'excel')
            print(f"🔍 Excel 디렉토리: {excel_dir}")
            if os.path.exists(excel_dir):
                print(f"🔍 Excel 디렉토리 내용: {os.listdir(excel_dir)}")
                
                # 카테고리별 디렉토리 확인
                if '/' in file_path:
                    category = file_path.split('/')[0]
                    category_path = os.path.join(excel_dir, category)
                    if os.path.exists(category_path):
                        print(f"🔍 {category} 디렉토리 내용: {os.listdir(category_path)}")
            
            if not os.path.exists(full_path):
                self._send_json_response({
                    "error": "파일을 찾을 수 없습니다",
                    "path": full_path,
                    "requested_path": file_path
                }, 404)
                return
            
            # Excel 파일 읽기
            print("📖 Excel 파일 읽기 시작...")
            workbook = openpyxl.load_workbook(full_path)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                print(f"📄 시트 처리 중: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # 시트 데이터를 리스트로 변환 (최대 100행까지만)
                data = []
                max_rows = min(sheet.max_row or 1, 100)
                max_cols = min(sheet.max_column or 1, 20)
                
                for row in range(1, max_rows + 1):
                    row_data = []
                    for col in range(1, max_cols + 1):
                        try:
                            cell_value = sheet.cell(row=row, column=col).value
                            row_data.append(str(cell_value) if cell_value is not None else "")
                        except Exception as cell_error:
                            print(f"⚠️ 셀 읽기 오류 ({row}, {col}): {cell_error}")
                            row_data.append("")
                    data.append(row_data)
                
                sheets_data[sheet_name] = {
                    'data': data,
                    'total_rows': sheet.max_row or 0,
                    'total_cols': sheet.max_column or 0
                }
            
            response_data = {
                'filename': os.path.basename(file_path),
                'sheets': sheets_data,
                'sheet_names': workbook.sheetnames
            }
            
            self._send_json_response(response_data)
            print(f"✅ 파일 미리보기 완료: {file_path}")
            
        except Exception as e:
            print(f"❌ 파일 미리보기 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            self._send_json_response({
                "error": f"파일 미리보기 실패: {str(e)}",
                "details": str(e)
            }, 500)

    def _handle_file_upload(self, post_data):
        """📤 파일 업로드"""
        try:
            print("📤 파일 업로드 요청 처리 중...")
            
            # JSON 데이터 파싱
            data = json.loads(post_data.decode('utf-8'))
            
            filename = data.get('filename')
            category = data.get('category')
            file_content = data.get('content')  # base64 encoded
            
            if not all([filename, category, file_content]):
                self._send_json_response({"error": "필수 파라미터가 누락되었습니다"}, 400)
                return
            
            # 파일 저장 경로
            script_dir = os.path.dirname(os.path.abspath(__file__))
            save_dir = os.path.join(script_dir, 'app', 'data', 'excel', category)
            os.makedirs(save_dir, exist_ok=True)
            
            save_path = os.path.join(save_dir, filename)
            
            # base64 디코딩 후 파일 저장
            file_bytes = base64.b64decode(file_content)
            with open(save_path, 'wb') as f:
                f.write(file_bytes)
            
            response_data = {
                'success': True,
                'message': '파일이 성공적으로 업로드되었습니다',
                'filename': filename,
                'category': category,
                'path': f"{category}/{filename}"
            }
            
            self._send_json_response(response_data)
            print(f"✅ 파일 업로드 완료: {filename}")
            
        except Exception as e:
            print(f"❌ 파일 업로드 오류: {str(e)}")
            self._send_json_response({"error": f"파일 업로드 실패: {str(e)}"}, 500)

    def _handle_file_save(self, file_path, post_data):
        """💾 파일 저장 (수정된 데이터)"""
        try:
            print(f"💾 파일 저장 요청: {file_path}")
            
            data = json.loads(post_data.decode('utf-8'))
            
            sheets_data = data.get('sheets')
            if not sheets_data:
                self._send_json_response({"error": "저장할 데이터가 없습니다"}, 400)
                return
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            full_path = os.path.join(script_dir, 'app', 'data', 'excel', file_path)
            
            # 기존 파일 백업
            backup_path = full_path + f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            if os.path.exists(full_path):
                shutil.copy2(full_path, backup_path)
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # 각 시트 데이터 저장
            for sheet_name, sheet_data in sheets_data.items():
                ws = workbook.create_sheet(title=sheet_name)
                
                for row_idx, row_data in enumerate(sheet_data['data'], 1):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # 파일 저장
            workbook.save(full_path)
            
            response_data = {
                'success': True,
                'message': '파일이 성공적으로 저장되었습니다',
                'backup_created': os.path.basename(backup_path) if os.path.exists(full_path) else None
            }
            
            self._send_json_response(response_data)
            print(f"✅ 파일 저장 완료: {file_path}")
            
        except Exception as e:
            print(f"❌ 파일 저장 오류: {str(e)}")
            self._send_json_response({"error": f"파일 저장 실패: {str(e)}"}, 500)

    def do_OPTIONS(self):
        """OPTIONS 요청 처리 (CORS preflight)"""
        self.send_response(200)
        self._set_cors_headers()
        self.end_headers()
    
    def do_GET(self):
        """GET 요청 처리"""
        parsed_path = urlparse(self.path)
        path = parsed_path.path
        
        print(f"🌐 GET 요청: {path}")
        print(f"🔍 EXCEL_AVAILABLE: {EXCEL_AVAILABLE}")
        
        # 루트 경로
        if path == '/':
            self._send_json_response({
                "message": "🏋️ Gym AI 백엔드 서버 실행 중!",
                "status": "success",
                "version": "basic-http-1.0",
                "excel_available": EXCEL_AVAILABLE
            })
            return
        
        # 인증 관련
        if path == '/api/v1/auth/me':
            self._send_json_response({
                "username": "admin",
                "email": "admin@gym.com",
                "id": 1
            })
            return
        
        # 회원 관리 API
        if path == '/api/v1/members/':
            print("📊 회원 데이터 요청 처리 중...")
            if EXCEL_AVAILABLE:
                try:
                    print("📋 Excel에서 회원 데이터 읽기 시도...")
                    members_data, summary = read_members_data()
                    print(f"✅ 읽은 회원 데이터: {len(members_data)}명, 요약: {summary}")
                    
                    response_data = {
                        "total_count": len(members_data),
                        "members": members_data,
                        "summary": summary,
                        "data_source": "Excel 파일 (분류형)",
                        "last_updated": "실시간"
                    }
                    print(f"📤 회원 데이터 응답 전송: {len(members_data)}명")
                    self._send_json_response(response_data)
                    return
                    
                except Exception as e:
                    print(f"❌ 회원 Excel 데이터 읽기 오류: {e}")
                    import traceback
                    traceback.print_exc()
                    self._send_json_response({
                        "error": "회원 Excel 데이터 읽기 실패",
                        "message": str(e)
                    }, 500)
                    return
            else:
                # Excel 모듈 없을 때 샘플 데이터
                members_data = [
                    {
                        "id": 1,
                        "name": "김철수",
                        "phone": "010-1234-5678",
                        "email": "kim@example.com",
                        "membership_type": "프리미엄",
                        "start_date": "2024-01-01",
                        "end_date": "2024-12-31",
                        "payment_status": "paid"
                    }
                ]
                
                self._send_json_response({
                    "total_count": len(members_data),
                    "members": members_data,
                    "summary": {"총회원수": len(members_data)},
                    "data_source": "샘플 데이터"
                })
                return
        
        # 직원 관리 API
        if path == '/api/v1/staff/':
            print("👥 직원 데이터 요청 처리 중...")
            if EXCEL_AVAILABLE:
                try:
                    print("📋 Excel에서 직원 데이터 읽기 시도...")
                    staff_data, summary = read_staff_data()
                    print(f"✅ 읽은 직원 데이터: {len(staff_data)}명, 요약: {summary}")
                    
                    response_data = {
                        "total_count": len(staff_data),
                        "staff": staff_data,
                        "summary": summary,
                        "data_source": "Excel 파일 (분류형)",
                        "last_updated": "실시간"
                    }
                    print(f"📤 직원 데이터 응답 전송: {len(staff_data)}명")
                    self._send_json_response(response_data)
                    return
                    
                except Exception as e:
                    print(f"❌ 직원 Excel 데이터 읽기 오류: {e}")
                    import traceback
                    traceback.print_exc()
                    self._send_json_response({
                        "error": "직원 Excel 데이터 읽기 실패",
                        "message": str(e)
                    }, 500)
                    return
            else:
                # Excel 모듈 없을 때 샘플 데이터
                staff_data = [
                    {
                        "id": 1,
                        "name": "최트레이너",
                        "phone": "010-1111-2222",
                        "email": "trainer@gym.com",
                        "position": "trainer",
                        "status": "active"
                    }
                ]
                
                self._send_json_response({
                    "total_count": len(staff_data),
                    "staff": staff_data,
                    "summary": {"총직원수": len(staff_data)},
                    "data_source": "샘플 데이터"
                })
                return
        
        # 재고 관리 API
        if path == '/api/v1/inventory/low-stock':
            print("📦 재고 데이터 요청 처리 중...")
            if EXCEL_AVAILABLE:
                try:
                    print("📋 Excel에서 재고 데이터 읽기 시도...")
                    inventory_data, summary, low_stock_data = read_inventory_data()
                    print(f"✅ 읽은 재고 데이터: 전체 {len(inventory_data)}개, 부족 {len(low_stock_data)}개")
                    
                    response_data = {
                        "low_stock_items": low_stock_data,
                        "alert_count": len(low_stock_data),
                        "inventory_summary": summary,
                        "data_source": "Excel 파일 (분류형)",
                        "last_updated": "실시간"
                    }
                    print(f"📤 재고 데이터 응답 전송: 부족재고 {len(low_stock_data)}개")
                    self._send_json_response(response_data)
                    return
                    
                except Exception as e:
                    print(f"❌ 재고 Excel 데이터 읽기 오류: {e}")
                    import traceback
                    traceback.print_exc()
                    self._send_json_response({
                        "error": "재고 Excel 데이터 읽기 실패",
                        "message": str(e)
                    }, 500)
                    return
            else:
                # Excel 모듈 없을 때 샘플 데이터
                low_stock_data = [
                    {
                        "id": 3,
                        "item_name": "운동 타올",
                        "current_stock": 15,
                        "min_stock_level": 20,
                        "category": "accessory",
                        "status": "부족"
                    }
                ]
                
                self._send_json_response({
                    "low_stock_items": low_stock_data,
                    "alert_count": len(low_stock_data),
                    "data_source": "샘플 데이터"
                })
                return
        
        # 📁 파일 관리 API
        if path == '/api/v1/files':
            self._handle_files_list()
            return
        elif path == '/api/v1/files/debug':
            # 디버깅용 API 추가
            script_dir = os.path.dirname(os.path.abspath(__file__))
            excel_dir = os.path.join(script_dir, 'app', 'data', 'excel')
            members_dir = os.path.join(excel_dir, 'members')
            
            debug_info = {
                "script_dir": script_dir,
                "excel_dir": excel_dir,
                "members_dir": members_dir,
                "excel_dir_exists": os.path.exists(excel_dir),
                "members_dir_exists": os.path.exists(members_dir),
                "members_files": os.listdir(members_dir) if os.path.exists(members_dir) else []
            }
            self._send_json_response(debug_info)
            return
        elif path.startswith('/api/v1/files/download/'):
            file_path = path.replace('/api/v1/files/download/', '')
            # URL 디코딩 추가
            from urllib.parse import unquote
            file_path = unquote(file_path)
            self._handle_file_download(file_path)
            return
        elif path.startswith('/api/v1/files/preview/'):
            file_path = path.replace('/api/v1/files/preview/', '')
            # URL 디코딩 추가
            from urllib.parse import unquote
            file_path = unquote(file_path)
            self._handle_file_preview(file_path)
            return

        # 대시보드 API
        if path == '/api/v1/dashboard':
            print("📊 대시보드 데이터 요청 처리 중...")
            if EXCEL_AVAILABLE:
                try:
                    print("📋 Excel에서 대시보드 데이터 읽기 시도...")
                    dashboard_data = get_all_dashboard_data()
                    print(f"✅ 대시보드 데이터 생성 완료: {dashboard_data}")
                    
                    response_data = {
                        "dashboard": dashboard_data,
                        "data_source": "Excel 파일 (분류형)",
                        "last_updated": "실시간"
                    }
                    self._send_json_response(response_data)
                    return
                    
                except Exception as e:
                    print(f"❌ 대시보드 데이터 읽기 오류: {e}")
                    import traceback
                    traceback.print_exc()
                    self._send_json_response({
                        "error": "대시보드 데이터 읽기 실패",
                        "message": str(e)
                    }, 500)
                    return
            else:
                # Excel 모듈 없을 때 기본 대시보드 데이터
                self._send_json_response({
                    "dashboard": {
                        "summary": {
                            "총회원수": 1,
                            "총직원수": 1,
                            "총품목수": 1,
                            "부족재고": 1
                        }
                    },
                    "data_source": "샘플 데이터"
                })
                return
        
        # 기타 API 경로 (fallback)
        if path.startswith('/api/v1/'):
            self._send_json_response({
                "message": "API 엔드포인트 정상 작동",
                "path": path,
                "method": "GET",
                "note": "이 메시지가 보이면 해당 경로가 구현되지 않았습니다."
            })
            return
        
        # 404 에러
        self._send_json_response({
            "error": "경로를 찾을 수 없습니다",
            "path": path
        }, 404)
    
    def do_POST(self):
        """POST 요청 처리"""
        parsed_path = urlparse(self.path)
        path = parsed_path.path
        
        print(f"🔐 POST 요청: {path}")
        
        # 요청 본문 읽기
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        
        try:
            if path == '/api/v1/auth/login':
                print("🚀 로그인 요청 처리 중...")
                
                # 로그인 성공 응답 (모든 계정 허용)
                response_data = {
                    "access_token": "basic_token_12345",
                    "token_type": "bearer", 
                    "user": {
                        "username": "admin",
                        "email": "admin@gym.com",
                        "id": 1
                    }
                }
                
                print("✅ 로그인 성공 응답 전송")
                self._send_json_response(response_data)
                return
            
            # 채팅 API 엔드포인트들
            if path == '/api/v1/members/chat':
                return self._handle_chat_request('회원관리', post_data)
            elif path == '/api/v1/staff/chat':
                return self._handle_chat_request('직원관리', post_data)
            elif path == '/api/v1/hr/chat':
                return self._handle_chat_request('인사관리', post_data)
            elif path == '/api/v1/inventory/chat':
                return self._handle_chat_request('재고관리', post_data)
            
            # 📁 파일 관리 API
            elif path == '/api/v1/files/upload':
                return self._handle_file_upload(post_data)
            elif path.startswith('/api/v1/files/save/'):
                file_path = path.replace('/api/v1/files/save/', '')
                # URL 디코딩 추가
                from urllib.parse import unquote
                file_path = unquote(file_path)
                return self._handle_file_save(file_path, post_data)
            
            # 기타 POST 요청
            self._send_json_response({
                "message": "POST 요청 처리됨",
                "path": path,
                "received_data_length": len(post_data)
            })
            
        except Exception as e:
            print(f"❌ POST 요청 처리 오류: {e}")
            self._send_json_response({
                "error": "요청 처리 중 오류 발생",
                "details": str(e)
            }, 500)

def run_server(port=8000):
    """서버 실행"""
    server_address = ('', port)
    httpd = HTTPServer(server_address, APIHandler)
    
    print("=" * 60)
    print("🚀 Gym AI 기본 HTTP 백엔드 서버 시작!")
    print(f"📍 서버 주소: http://localhost:{port}")
    print(f"🔐 로그인 API: http://localhost:{port}/api/v1/auth/login")
    print("💡 모든 사용자명/비밀번호로 로그인 가능!")
    print(f"📊 Excel 모듈 상태: {'✅ 사용 가능' if EXCEL_AVAILABLE else '❌ 사용 불가'}")
    print("=" * 60)
    
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n🛑 서버 종료 중...")
        httpd.shutdown()

if __name__ == "__main__":
    run_server(8000) 